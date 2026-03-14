import asyncio
import io
import logging
import socket
import sys
import uuid

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import delete as sa_delete
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from cctv_monitor.api.deps import get_session, get_settings, get_driver_registry, get_http_client
from pydantic import BaseModel
from cctv_monitor.api.schemas import (
    DeviceCreate, DeviceUpdate, DeviceOut, DeviceDetailOut, PollResultOut,
    HealthSummaryOut, AlertOut, CameraChannelOut, DiskOut,
)
from cctv_monitor.core.config import Settings
from cctv_monitor.core.crypto import encrypt_value, decrypt_value
from cctv_monitor.core.types import DeviceTransport, DeviceVendor
from cctv_monitor.drivers.hikvision.errors import IsapiAuthError, IsapiError, SdkError
from cctv_monitor.drivers.hikvision.transports.isapi import IsapiTransport
from cctv_monitor.drivers.registry import DriverRegistry
from cctv_monitor.core.http_client import HttpClientManager
from cctv_monitor.models.device import DeviceConfig
from cctv_monitor.storage.repositories import DeviceRepository, AlertRepository, DeviceTagRepository, DeviceHealthLogRepository, FolderRepository
from cctv_monitor.storage.tables import AlertTable, DeviceTable

router = APIRouter(tags=["devices"])


async def _check_tcp_port(host: str, port: int, timeout: float = 3.0) -> bool:
    """Check if a TCP port is open (non-blocking)."""
    loop = asyncio.get_event_loop()
    try:
        def _probe():
            with socket.create_connection((host, port), timeout=timeout):
                pass
        await loop.run_in_executor(None, _probe)
        return True
    except (OSError, TimeoutError):
        return False


def _device_out(d: DeviceTable, tags: list[dict] | None = None, folder_path: str | None = None) -> DeviceOut:
    cached = d.last_health_json or {}
    health_data = cached.get("health")
    health = None
    if health_data:
        health = HealthSummaryOut(**health_data)
        # Adjust counts: exclude ignored channels
        ignored = {str(ch) for ch in (d.ignored_channels or [])}
        if ignored:
            cameras_data = cached.get("cameras", [])
            monitored = [c for c in cameras_data if str(c.get("channel_id", "")) not in ignored]
            monitored_online = sum(
                1 for c in monitored if c.get("status", "").lower() == "online"
            ) if health.reachable else 0
            health.camera_count = len(monitored)
            health.online_cameras = monitored_online
            health.offline_cameras = len(monitored) - monitored_online if health.reachable else 0
    return DeviceOut(
        device_id=d.device_id, name=d.name, vendor=d.vendor,
        host=d.host, web_port=d.web_port, sdk_port=d.sdk_port,
        web_protocol=d.web_protocol,
        transport_mode=d.transport_mode, is_active=d.is_active,
        last_health=health,
        model=d.model, serial_number=d.serial_number,
        firmware_version=d.firmware_version, last_poll_at=d.last_poll_at,
        poll_interval_seconds=d.poll_interval_seconds,
        tags=tags or [],
        ignored_channels=d.ignored_channels or [],
        folder_id=d.folder_id,
        folder_path=folder_path,
        display_order=d.display_order,
    )


async def _build_folder_paths(session: AsyncSession) -> dict[int, str]:
    """Build a map of folder_id -> display path (e.g. 'Company / Branch')."""
    folder_repo = FolderRepository(session)
    folders = await folder_repo.list_all()
    folder_map = {f.id: f for f in folders}
    paths: dict[int, str] = {}
    for f in folders:
        if f.parent_id is None:
            paths[f.id] = f.name
        else:
            parent = folder_map.get(f.parent_id)
            paths[f.id] = f"{parent.name} / {f.name}" if parent else f.name
    return paths


@router.get("/devices", response_model=list[DeviceOut])
async def list_devices(
    tag: str | None = None,
    search: str | None = None,
    folder_id: int | None = None,
    session: AsyncSession = Depends(get_session),
):
    repo = DeviceRepository(session)
    devices = await repo.list_all()
    tag_repo = DeviceTagRepository(session)
    folder_paths = await _build_folder_paths(session)

    result = []
    for d in devices:
        if folder_id is not None and d.folder_id != folder_id:
            continue
        tags = await tag_repo.get_tags_with_colors(d.device_id)
        tag_names = [t["name"] for t in tags]
        if tag and tag not in tag_names:
            continue
        if search and search.lower() not in d.name.lower() and search not in d.host:
            continue
        path = folder_paths.get(d.folder_id) if d.folder_id else None
        result.append(_device_out(d, tags, folder_path=path))
    return result


@router.get("/devices/export")
async def export_devices_excel(
    folder_ids: str | None = Query(None, description="Comma-separated folder IDs, empty = all"),
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    repo = DeviceRepository(session)
    folder_repo = FolderRepository(session)
    all_devices = await repo.list_all()
    folders = await folder_repo.list_all()

    # Parse folder filter
    selected_ids: set[int] | None = None
    if folder_ids:
        selected_ids = set()
        for raw in folder_ids.split(","):
            raw = raw.strip()
            if raw.isdigit():
                selected_ids.add(int(raw))

    # Build folder hierarchy
    folder_map = {f.id: f for f in folders}
    top_folders = [f for f in folders if f.parent_id is None]
    children_map: dict[int, list] = {}
    for f in folders:
        if f.parent_id is not None:
            children_map.setdefault(f.parent_id, []).append(f)

    # Group devices by folder
    devices_by_folder: dict[int | None, list[DeviceTable]] = {}
    for d in all_devices:
        devices_by_folder.setdefault(d.folder_id, []).append(d)

    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"
    ws.sheet_view.rightToLeft = True  # RTL for Hebrew

    # Styles
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    folder_font = Font(name="Arial", bold=True, size=11)
    data_font = Font(name="Arial", size=10)
    data_alignment = Alignment(horizontal="right", vertical="center")
    mono_font = Font(name="Consolas", size=10)

    columns = ["שם", "כתובת IP", "פורט Web", "פורט SDK", "שם משתמש", "סיסמה"]
    col_widths = [35, 18, 12, 12, 16, 18]

    # Set column widths
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    def write_header(ws, row: int):
        for col, title in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        return row + 1

    def write_folder_row(ws, row: int, label: str, color_hex: str | None):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        bg = (color_hex or "#3B82F6").lstrip("#")
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = thin_border
        # Apply border to all merged cells
        for col in range(2, len(columns) + 1):
            ws.cell(row=row, column=col).border = thin_border
        return row + 1

    def write_device_row(ws, row: int, d: DeviceTable, stripe: bool):
        password = ""
        try:
            password = decrypt_value(d.password_encrypted, settings.ENCRYPTION_KEY)
        except Exception:
            password = "***"
        values = [d.name, d.host, d.web_port or "", d.sdk_port or "", d.username, password]
        bg_color = "F2F2F2" if stripe else "FFFFFF"
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = mono_font if col == 2 else data_font
            cell.alignment = data_alignment
            cell.fill = fill
            cell.border = thin_border
        return row + 1

    def process_folder(ws, row: int, folder, level: int = 0) -> int:
        folder_ids_set = {folder.id}
        kids = children_map.get(folder.id, [])
        for kid in kids:
            folder_ids_set.add(kid.id)

        # Check if this folder (or children) should be included
        if selected_ids is not None:
            if not folder_ids_set.intersection(selected_ids):
                return row

        # Folder header
        row = write_folder_row(ws, row, ("  " * level) + folder.name, folder.color)

        # Direct devices
        devs = devices_by_folder.get(folder.id, [])
        for i, d in enumerate(devs):
            row = write_device_row(ws, row, d, i % 2 == 0)

        # Subfolders
        for kid in kids:
            if selected_ids is not None and kid.id not in selected_ids and folder.id not in selected_ids:
                continue
            kid_devs = devices_by_folder.get(kid.id, [])
            if kid_devs or (selected_ids is not None and kid.id in selected_ids):
                row = write_folder_row(ws, row, "  " + kid.name, kid.color or folder.color)
                for i, d in enumerate(kid_devs):
                    row = write_device_row(ws, row, d, i % 2 == 0)

        return row

    # Header row
    row = write_header(ws, row)

    # Process each top folder
    for tf in top_folders:
        row = process_folder(ws, row, tf)

    # No-folder devices
    no_folder_devs = devices_by_folder.get(None, [])
    if no_folder_devs and (selected_ids is None or not selected_ids):
        row = write_folder_row(ws, row, "ללא תיקייה", "#9CA3AF")
        for i, d in enumerate(no_folder_devs):
            row = write_device_row(ws, row, d, i % 2 == 0)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=devices_export.xlsx"},
    )


@router.post("/devices", response_model=DeviceOut, status_code=201)
async def create_device(
    body: DeviceCreate,
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
):
    device_id = body.device_id or f"dvr-{uuid.uuid4().hex[:8]}"
    encrypted_password = encrypt_value(body.password, settings.ENCRYPTION_KEY)
    device = DeviceTable(
        device_id=device_id, name=body.name, vendor=body.vendor,
        host=body.host, web_port=body.web_port, sdk_port=body.sdk_port,
        web_protocol=body.web_protocol,
        username=body.username, password_encrypted=encrypted_password,
        transport_mode=body.transport_mode, is_active=True,
        poll_interval_seconds=body.poll_interval_seconds,
        folder_id=body.folder_id,
    )
    repo = DeviceRepository(session)
    try:
        await repo.create(device)
        await session.commit()
    except IntegrityError as exc:
        await session.rollback()
        err = str(exc.orig) if exc.orig else str(exc)
        if "polling_policies" in err or "foreign key" in err.lower():
            raise HTTPException(status_code=500, detail="Polling policy 'standard' not found. Run seed first.")
        raise HTTPException(status_code=409, detail=f"Device '{body.device_id}' already exists")
    return _device_out(device)


# ---------- Bulk reorder ----------

class _ReorderItem(BaseModel):
    device_id: str
    display_order: int
    folder_id: int | None = None


class _ReorderRequest(BaseModel):
    items: list[_ReorderItem]


@router.put("/devices/reorder")
async def reorder_devices(
    body: _ReorderRequest,
    session: AsyncSession = Depends(get_session),
):
    repo = DeviceRepository(session)
    for item in body.items:
        device = await repo.get_by_id(item.device_id)
        if device:
            device.display_order = item.display_order
            if item.folder_id is not None:
                device.folder_id = item.folder_id
    await session.commit()
    return {"ok": True}


@router.delete("/devices/{device_id}", status_code=204)
async def delete_device(device_id: str, session: AsyncSession = Depends(get_session)):
    repo = DeviceRepository(session)
    # Delete related alerts first to avoid FK constraint violation
    await session.execute(sa_delete(AlertTable).where(AlertTable.device_id == device_id))
    deleted = await repo.delete(device_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Device not found")
    await session.commit()


@router.patch("/devices/{device_id}", response_model=DeviceOut)
async def update_device(
    device_id: str,
    body: DeviceUpdate,
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
):
    repo = DeviceRepository(session)
    fields = body.model_dump(exclude_unset=True)
    if "password" in fields:
        fields["password_encrypted"] = encrypt_value(fields.pop("password"), settings.ENCRYPTION_KEY)
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    device = await repo.update(device_id, **fields)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")
    await session.commit()
    return _device_out(device)


@router.get("/devices/{device_id}", response_model=DeviceDetailOut)
async def get_device_detail(device_id: str, session: AsyncSession = Depends(get_session)):
    repo = DeviceRepository(session)
    device = await repo.get_by_id(device_id)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")

    tag_repo = DeviceTagRepository(session)
    tags = await tag_repo.get_tags_with_colors(device_id)

    alert_repo = AlertRepository(session)
    alerts = await alert_repo.get_active_alerts(device_id)

    # Extract from cached last_health_json
    cached = device.last_health_json or {}
    cameras = [CameraChannelOut(**c) for c in cached.get("cameras", [])]
    disks = [DiskOut(**d) for d in cached.get("disks", [])]
    health_data = cached.get("health")
    health = HealthSummaryOut(**health_data) if health_data else None

    return DeviceDetailOut(
        device=_device_out(device, tags),
        cameras=cameras, disks=disks,
        alerts=[
            AlertOut(
                id=a.id, device_id=a.device_id, device_name=device.name,
                alert_type=a.alert_type, severity=a.severity,
                message=a.message, status=a.status, created_at=a.created_at,
                resolved_at=a.resolved_at,
            ) for a in alerts
        ],
        health=health,
    )


@router.get("/devices/{device_id}/credentials")
async def get_credentials(
    device_id: str,
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
):
    repo = DeviceRepository(session)
    device = await repo.get_by_id(device_id)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")
    password = decrypt_value(device.password_encrypted, settings.ENCRYPTION_KEY)
    return {"username": device.username, "password": password}


@router.get("/devices/{device_id}/ignored-channels", response_model=list[str])
async def get_ignored_channels(device_id: str, session: AsyncSession = Depends(get_session)):
    repo = DeviceRepository(session)
    device = await repo.get_by_id(device_id)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")
    return device.ignored_channels or []


@router.put("/devices/{device_id}/ignored-channels", response_model=list[str])
async def set_ignored_channels(
    device_id: str,
    channels: list[str],
    session: AsyncSession = Depends(get_session),
):
    repo = DeviceRepository(session)
    device = await repo.get_by_id(device_id)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")
    device.ignored_channels = channels
    await session.commit()
    return device.ignored_channels or []


def _build_time_sync_xml(iso_time: str, timezone: str) -> str:
    """Build XML body for ISAPI time sync PUT."""
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Time>"
        "<timeMode>manual</timeMode>"
        f"<localTime>{iso_time}</localTime>"
        f"<timeZone>{timezone}</timeZone>"
        "</Time>"
    )


def _get_israel_time() -> tuple[str, str]:
    """Return (iso_time, tz_posix) for current Israel time."""
    from datetime import datetime, timezone as tz, timedelta

    israel_tz_posix = "CST-2:00:00DST01:00:00,M3.5.5/02:00:00,M10.5.0/02:00:00"
    now_utc = datetime.now(tz.utc)
    try:
        from zoneinfo import ZoneInfo
        now_israel = now_utc.astimezone(ZoneInfo("Asia/Jerusalem"))
    except Exception:
        now_israel = now_utc.astimezone(tz(timedelta(hours=2)))

    iso_time = (
        now_israel.strftime("%Y-%m-%dT%H:%M:%S")
        + now_israel.strftime("%z")[:3] + ":" + now_israel.strftime("%z")[3:]
    )
    return iso_time, israel_tz_posix


async def _sync_time_via_sdk(
    host: str, sdk_port: int, username: str, password: str,
    xml_body: str, lib_path: str,
) -> dict:
    """Sync time via SDK subprocess ISAPI tunnel."""
    import asyncio
    cmd = [
        sys.executable, "-m", "cctv_monitor.polling.sdk_worker",
        "--host", host,
        "--port", str(sdk_port),
        "--user", username,
        "--password", password,
        "--lib-path", lib_path,
        "--sync-time",
        "--time-xml", xml_body,
    ]
    proc = await asyncio.create_subprocess_exec(
        *cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=30)
    if proc.returncode != 0:
        raise RuntimeError(f"SDK worker exited {proc.returncode}: {stderr.decode(errors='ignore')}")
    import json as _json
    result = _json.loads(stdout.decode(errors="ignore"))
    if not result.get("success"):
        raise RuntimeError(result.get("error", "SDK time sync failed"))
    return result


@router.post("/devices/{device_id}/sync-time")
async def sync_device_time(
    device_id: str,
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
    http_client: HttpClientManager = Depends(get_http_client),
):
    """Sync device time to current server time (Israel timezone) via ISAPI or SDK."""
    repo = DeviceRepository(session)
    device = await repo.get_by_id(device_id)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")

    if not device.web_port and not (device.sdk_port and settings.HCNETSDK_LIB_PATH):
        raise HTTPException(status_code=400, detail="No web port or SDK port configured")

    password = decrypt_value(device.password_encrypted, settings.ENCRYPTION_KEY)
    iso_time, israel_tz_posix = _get_israel_time()
    xml_body = _build_time_sync_xml(iso_time, israel_tz_posix)
    sdk_available = bool(device.sdk_port and settings.HCNETSDK_LIB_PATH)

    # Try ISAPI first if web_port is configured
    if device.web_port:
        try:
            transport = IsapiTransport(http_client)
            await transport.connect(device.host, device.web_port, device.username, password, protocol=device.web_protocol)
            try:
                result = await transport.set_device_time(iso_time, israel_tz_posix)
                return {
                    "success": result["status_code"] < 300,
                    "time_set": iso_time,
                    "timezone": israel_tz_posix,
                    "transport": "isapi",
                    "status_code": result["status_code"],
                }
            finally:
                try:
                    await transport.disconnect()
                except Exception:
                    pass
        except IsapiAuthError as exc:
            raise HTTPException(status_code=502, detail=f"Auth failed: {exc}")
        except Exception as exc:
            if not sdk_available:
                raise HTTPException(status_code=502, detail=f"Time sync failed: {exc}")
            logger.info("ISAPI time sync failed for device=%s, falling back to SDK: %s", device_id, exc)

    # Fallback to SDK ISAPI tunnel
    if sdk_available:
        try:
            await _sync_time_via_sdk(
                device.host, device.sdk_port, device.username, password,
                xml_body, settings.HCNETSDK_LIB_PATH,
            )
            return {
                "success": True,
                "time_set": iso_time,
                "timezone": israel_tz_posix,
                "transport": "sdk",
            }
        except Exception as exc:
            logger.warning("SDK time sync failed for device=%s: %s", device_id, exc)
            raise HTTPException(status_code=502, detail=f"Time sync failed: {exc}")

    raise HTTPException(status_code=400, detail="No web port or SDK port configured")


def _parse_network_interfaces_xml(raw_xml: str) -> list[dict]:
    """Parse ISAPI /System/Network/interfaces XML into interface dicts."""
    import xml.etree.ElementTree as ET
    interfaces = []
    try:
        root = ET.fromstring(raw_xml)
    except ET.ParseError:
        return interfaces
    ns = ""
    if root.tag.startswith("{"):
        ns = root.tag.split("}")[0] + "}"
    for iface in root.iter(f"{ns}NetworkInterface"):
        iface_id = iface.findtext(f"{ns}id", "")
        # Collect ALL text values from the subtree for robust parsing
        def _find_text(*paths: str) -> str | None:
            for p in paths:
                el = iface.find(p)
                if el is not None and el.text:
                    return el.text.strip()
            return None
        ip = _find_text(
            f".//{ns}IPAddress/{ns}ipAddress",
            f".//{ns}IPv4Addressing/{ns}ipAddress",
            f".//{ns}ipAddress",
        )
        mask = _find_text(
            f".//{ns}IPAddress/{ns}subnetMask",
            f".//{ns}IPv4Addressing/{ns}subnetMask",
            f".//{ns}subnetMask",
        )
        gateway = _find_text(
            f".//{ns}IPAddress/{ns}DefaultGateway/{ns}ipAddress",
            f".//{ns}DefaultGateway/{ns}ipAddress",
            f".//{ns}defaultGateway/{ns}ipAddress",
            f".//{ns}IPAddress/{ns}ipv4DefaultGateway",
            f".//{ns}ipv4DefaultGateway",
        )
        mac = _find_text(
            f".//{ns}Link/{ns}MACAddress",
            f".//{ns}MACAddress",
            f".//{ns}macAddress",
        )
        # DNS
        dns1 = _find_text(
            f".//{ns}IPAddress/{ns}PrimaryDNS/{ns}ipAddress",
            f".//{ns}PrimaryDNS/{ns}ipAddress",
            f".//{ns}primaryDNS/{ns}ipAddress",
        )
        dns2 = _find_text(
            f".//{ns}IPAddress/{ns}SecondaryDNS/{ns}ipAddress",
            f".//{ns}SecondaryDNS/{ns}ipAddress",
            f".//{ns}secondaryDNS/{ns}ipAddress",
        )
        interfaces.append({
            "id": iface_id,
            "ip": ip,
            "mask": mask,
            "gateway": gateway,
            "mac": mac,
            "dns1": dns1,
            "dns2": dns2,
        })
    return interfaces


def _parse_network_ports_xml(raw_xml: str) -> list[dict]:
    """Parse ISAPI /Security/adminAccesses XML into port dicts."""
    import xml.etree.ElementTree as ET
    ports = []
    try:
        root = ET.fromstring(raw_xml)
    except ET.ParseError:
        return ports
    ns = ""
    if root.tag.startswith("{"):
        ns = root.tag.split("}")[0] + "}"
    for access in root.iter(f"{ns}AdminAccessProtocol"):
        proto = access.findtext(f"{ns}protocol", "")
        port_num = access.findtext(f"{ns}portNo", "")
        enabled = access.findtext(f"{ns}enabled", "true")
        if proto and port_num:
            ports.append({
                "protocol": proto,
                "port": int(port_num),
                "enabled": enabled.lower() == "true",
            })
    return ports


async def _get_network_via_sdk(host: str, sdk_port: int, username: str, password: str, lib_path: str) -> tuple[list[dict], list[dict]]:
    """Get network info via SDK ISAPI tunnel (subprocess)."""
    import json as _json
    interfaces: list[dict] = []
    ports: list[dict] = []

    cmd = [
        sys.executable, "-m", "cctv_monitor.polling.sdk_worker",
        "--host", host, "--port", str(sdk_port),
        "--user", username, "--password", password,
        "--lib-path", lib_path,
        "--isapi-get", "GET /ISAPI/System/Network/interfaces",
    ]
    try:
        proc = await asyncio.create_subprocess_exec(*cmd, stdout=PIPE, stderr=PIPE)
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=30)
        result = _json.loads(stdout.decode())
        if result.get("success") and result.get("response"):
            interfaces = _parse_network_interfaces_xml(result["response"])
            if not interfaces:
                logger.warning("SDK network interfaces: got XML but parsed 0 interfaces. XML[:500]=%s", result["response"][:500])
        elif result.get("error"):
            logger.warning("SDK network interfaces error: %s", result["error"])
        if stderr:
            logger.debug("SDK network interfaces stderr: %s", stderr.decode(errors="replace").strip())
    except Exception as exc:
        logger.warning("SDK network interfaces failed: %s", exc)

    cmd_ports = [
        sys.executable, "-m", "cctv_monitor.polling.sdk_worker",
        "--host", host, "--port", str(sdk_port),
        "--user", username, "--password", password,
        "--lib-path", lib_path,
        "--isapi-get", "GET /ISAPI/Security/adminAccesses",
    ]
    try:
        proc = await asyncio.create_subprocess_exec(*cmd_ports, stdout=PIPE, stderr=PIPE)
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=30)
        result = _json.loads(stdout.decode())
        if result.get("success") and result.get("response"):
            ports = _parse_network_ports_xml(result["response"])
            if not ports:
                logger.warning("SDK network ports: got XML but parsed 0 ports. XML[:500]=%s", result["response"][:500])
        elif result.get("error"):
            logger.warning("SDK network ports error: %s", result["error"])
        if stderr:
            logger.debug("SDK network ports stderr: %s", stderr.decode(errors="replace").strip())
    except Exception as exc:
        logger.warning("SDK network ports failed: %s", exc)

    return interfaces, ports


@router.get("/devices/{device_id}/network")
async def get_device_network(
    device_id: str,
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
    http_client: HttpClientManager = Depends(get_http_client),
):
    """Get device network configuration via ISAPI (HTTP or SDK tunnel)."""
    repo = DeviceRepository(session)
    device = await repo.get_by_id(device_id)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")

    password = decrypt_value(device.password_encrypted, settings.ENCRYPTION_KEY)
    interfaces: list[dict] = []
    ports: list[dict] = []
    isapi_tried = False

    # Try ISAPI (HTTP) first
    if device.web_port:
        isapi_tried = True
        transport = IsapiTransport(http_client)
        try:
            await transport.connect(device.host, device.web_port, device.username, password, protocol=device.web_protocol)

            try:
                net_result = await transport.get_network_interfaces()
                raw_xml = net_result.get("raw_xml", "")
                if raw_xml:
                    interfaces = _parse_network_interfaces_xml(raw_xml)
            except Exception as exc:
                logger.warning("Failed to get network interfaces for device=%s: %s", device_id, exc)

            try:
                ports_result = await transport.get_network_ports()
                raw_xml = ports_result.get("raw_xml", "")
                if raw_xml:
                    ports = _parse_network_ports_xml(raw_xml)
            except Exception as exc:
                logger.warning("Failed to get ports for device=%s: %s", device_id, exc)

        except IsapiAuthError as exc:
            raise HTTPException(status_code=502, detail=f"Auth failed: {exc}")
        except Exception as exc:
            logger.warning("ISAPI network failed for device=%s: %s", device_id, exc)
        finally:
            try:
                await transport.disconnect()
            except Exception:
                pass

    # Fallback to SDK ISAPI tunnel if ISAPI failed or not configured
    sdk_available = device.sdk_port and settings.HCNETSDK_LIB_PATH
    if not interfaces and not ports and sdk_available:
        logger.info("Trying SDK ISAPI tunnel for network info device=%s", device_id)
        interfaces, ports = await _get_network_via_sdk(
            device.host, device.sdk_port, device.username, password,
            settings.HCNETSDK_LIB_PATH,
        )

    if not interfaces and not ports and not isapi_tried and not sdk_available:
        raise HTTPException(status_code=400, detail="No web port or SDK port configured")

    return {"interfaces": interfaces, "ports": ports}


@router.post("/devices/{device_id}/poll", response_model=PollResultOut)
async def poll_device(
    device_id: str,
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
    registry: DriverRegistry = Depends(get_driver_registry),
    http_client: HttpClientManager = Depends(get_http_client),
):
    repo = DeviceRepository(session)
    device = await repo.get_by_id(device_id)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")

    password = decrypt_value(device.password_encrypted, settings.ENCRYPTION_KEY)

    # TCP-check both ports in parallel (for display only, not for transport selection)
    web_port_open = None
    sdk_port_open = None
    if device.web_port and device.sdk_port:
        web_port_open, sdk_port_open = await asyncio.gather(
            _check_tcp_port(device.host, device.web_port, timeout=5.0),
            _check_tcp_port(device.host, device.sdk_port, timeout=5.0),
        )
    elif device.web_port:
        web_port_open = await _check_tcp_port(device.host, device.web_port, timeout=5.0)
    elif device.sdk_port:
        sdk_port_open = await _check_tcp_port(device.host, device.sdk_port, timeout=5.0)

    # Strategy: try ISAPI first (if web_port open), then SDK (if sdk_port configured).
    # TCP check is advisory — we always try the transport if port is configured.

    # 1) Try ISAPI if web port is open
    if device.web_port and web_port_open:
        try:
            return await _poll_via_isapi(
                device, password, device_id, repo, session,
                registry, http_client,
                web_port_open=web_port_open, sdk_port_open=sdk_port_open,
            )
        except IsapiAuthError:
            # Auth failed — do NOT try SDK with same bad credentials
            logger.warning("poll.auth_failed device_id=%s — skipping SDK fallback", device_id)
            from datetime import datetime, timezone as tz
            now = datetime.now(tz.utc)
            health = HealthSummaryOut(
                reachable=False, camera_count=0, online_cameras=0,
                offline_cameras=0, disk_ok=False, response_time_ms=0,
                checked_at=now,
                web_port_open=web_port_open, sdk_port_open=sdk_port_open,
            )
            raise HTTPException(
                status_code=502,
                detail="Authentication failed — wrong username or password. SDK fallback skipped to avoid device lockout.",
            )

    # 2) Try SDK subprocess if sdk_port is configured (regardless of TCP check result)
    if device.sdk_port and settings.HCNETSDK_LIB_PATH:
        result = await _poll_via_sdk_subprocess(
            device, password, device_id, repo, session,
            settings.HCNETSDK_LIB_PATH,
            web_port_open=web_port_open, sdk_port_open=sdk_port_open,
        )
        # If SDK succeeded, update sdk_port_open to True (it clearly works)
        if result.health.reachable and sdk_port_open is False:
            result.health.sdk_port_open = True
        return result

    # 3) No transport available — save failure
    from datetime import datetime, timezone as tz
    now = datetime.now(tz.utc)
    health = HealthSummaryOut(
        reachable=False, camera_count=0, online_cameras=0,
        offline_cameras=0, disk_ok=False, response_time_ms=0,
        checked_at=now,
        web_port_open=web_port_open, sdk_port_open=sdk_port_open,
    )
    health_json = {
        "reachable": False, "camera_count": 0,
        "online_cameras": 0, "offline_cameras": 0,
        "disk_ok": False, "response_time_ms": 0,
        "checked_at": now.isoformat(),
        "web_port_open": web_port_open, "sdk_port_open": sdk_port_open,
    }
    try:
        await repo.update(device_id, last_poll_at=now, last_health_json={
            "cameras": [], "disks": [], "health": health_json,
        })
        health_log_repo = DeviceHealthLogRepository(session)
        await health_log_repo.insert(
            device_id=device_id, reachable=False,
            camera_count=0, online_cameras=0,
            offline_cameras=0, disk_ok=False, response_time_ms=0,
        )
        await session.commit()
    except Exception:
        pass
    return PollResultOut(device_id=device_id, health=health)


async def _poll_via_isapi(
    device, password: str, device_id: str,
    repo: DeviceRepository, session: AsyncSession,
    registry: DriverRegistry, http_client: HttpClientManager,
    web_port_open: bool | None = None, sdk_port_open: bool | None = None,
) -> PollResultOut:
    """Poll device via ISAPI transport (in-process, safe)."""
    from datetime import datetime, timezone
    import time as _time

    vendor = DeviceVendor(device.vendor)
    driver_cls = registry.get(vendor)
    transport = IsapiTransport(http_client)
    driver = driver_cls(transport)

    config = DeviceConfig(
        device_id=device.device_id, name=device.name, vendor=vendor,
        host=device.host, web_port=device.web_port, sdk_port=device.sdk_port,
        web_protocol=device.web_protocol,
        username=device.username, password=password,
        transport_mode=DeviceTransport(device.transport_mode),
        polling_policy_id=device.polling_policy_id, is_active=device.is_active,
    )

    try:
        await driver.connect(config, port=device.web_port)
        start_t = _time.monotonic()

        device_info = None
        try:
            device_info = await driver.get_device_info()
        except Exception:
            pass

        cameras_raw = []
        try:
            cameras_raw = await driver.get_camera_statuses()
        except Exception:
            pass

        disks_raw = []
        try:
            disks_raw = await driver.get_disk_statuses()
        except Exception:
            pass

        # Recording status — prefer SDK FindFile (reliable) over ISAPI search
        rec_map: dict[str, str] = {}
        if device.sdk_port and settings.HCNETSDK_LIB_PATH:
            try:
                from cctv_monitor.polling.sdk_subprocess import poll_device_via_sdk_recordings
                sdk_recs = await poll_device_via_sdk_recordings(
                    host=device.host, port=device.sdk_port,
                    username=device.username, password=password,
                    lib_path=settings.HCNETSDK_LIB_PATH,
                    channels=[c.channel_id for c in cameras_raw],
                )
                for r in sdk_recs:
                    rec_map[r.get("channel_id", "")] = r.get("recording", "unknown")
            except Exception:
                pass
        if not rec_map:
            try:
                rec_statuses = await driver.get_recording_statuses()
                for r in rec_statuses:
                    rec_map[r.channel_id] = r.status.value
            except Exception:
                pass

        # Time check
        time_check_data: dict | None = None
        try:
            time_check_data = await driver.get_device_time()
        except Exception:
            pass

        response_time = (_time.monotonic() - start_t) * 1000
        now = datetime.now(timezone.utc)

        online = sum(1 for c in cameras_raw if c.status.value == "online")
        disk_ok = all(d.status.value == "ok" for d in disks_raw) if disks_raw else True

        health = HealthSummaryOut(
            reachable=True, camera_count=len(cameras_raw),
            online_cameras=online, offline_cameras=len(cameras_raw) - online,
            disk_ok=disk_ok, response_time_ms=response_time,
            checked_at=now,
            web_port_open=web_port_open, sdk_port_open=sdk_port_open,
        )

        cameras_json = [
            {"channel_id": c.channel_id, "channel_name": c.channel_name,
             "status": c.status.value, "ip_address": c.ip_address,
             "recording": rec_map.get(c.channel_id),
             "checked_at": c.checked_at.isoformat()}
            for c in cameras_raw
        ]
        disks_json = [
            {"disk_id": d.disk_id, "status": d.status.value,
             "capacity_bytes": d.capacity_bytes, "free_bytes": d.free_bytes,
             "health_status": d.health_status,
             "checked_at": d.checked_at.isoformat(),
             "temperature": d.temperature, "power_on_hours": d.power_on_hours,
             "smart_status": d.smart_status}
            for d in disks_raw
        ]
        health_json = {
            "reachable": True, "camera_count": len(cameras_raw),
            "online_cameras": online, "offline_cameras": len(cameras_raw) - online,
            "disk_ok": disk_ok, "response_time_ms": response_time,
            "checked_at": now.isoformat(),
            "web_port_open": web_port_open, "sdk_port_open": sdk_port_open,
            "time_check": time_check_data,
        }

        update_fields: dict = {
            "last_poll_at": now,
            "last_health_json": {
                "cameras": cameras_json, "disks": disks_json, "health": health_json,
            },
        }
        if device_info:
            update_fields["model"] = device_info.model
            update_fields["serial_number"] = device_info.serial_number
            update_fields["firmware_version"] = device_info.firmware_version

        await repo.update(device_id, **update_fields)

        health_log_repo = DeviceHealthLogRepository(session)
        await health_log_repo.insert(
            device_id=device_id, reachable=True,
            camera_count=len(cameras_raw), online_cameras=online,
            offline_cameras=len(cameras_raw) - online,
            disk_ok=disk_ok, response_time_ms=response_time,
        )
        await session.commit()

    except IsapiAuthError:
        # Re-raise auth errors so callers can avoid SDK fallback
        raise
    except Exception:
        now_err = datetime.now(timezone.utc)
        health = HealthSummaryOut(
            reachable=False, camera_count=0, online_cameras=0,
            offline_cameras=0, disk_ok=False, response_time_ms=0,
            checked_at=now_err,
            web_port_open=web_port_open, sdk_port_open=sdk_port_open,
        )
        health_json_err = {
            "reachable": False, "camera_count": 0,
            "online_cameras": 0, "offline_cameras": 0,
            "disk_ok": False, "response_time_ms": 0,
            "checked_at": now_err.isoformat(),
            "web_port_open": web_port_open, "sdk_port_open": sdk_port_open,
        }
        try:
            health_log_repo = DeviceHealthLogRepository(session)
            await health_log_repo.insert(
                device_id=device_id, reachable=False,
                camera_count=0, online_cameras=0,
                offline_cameras=0, disk_ok=False, response_time_ms=0,
            )
            await repo.update(device_id, last_poll_at=now_err, last_health_json={
                "cameras": [], "disks": [], "health": health_json_err,
            })
            await session.commit()
        except Exception:
            pass
        return PollResultOut(device_id=device_id, health=health)
    finally:
        try:
            await driver.disconnect()
        except Exception:
            pass

    return PollResultOut(device_id=device_id, health=health)


async def _poll_via_sdk_subprocess(
    device, password: str, device_id: str,
    repo: DeviceRepository, session: AsyncSession,
    lib_path: str,
    web_port_open: bool | None = None, sdk_port_open: bool | None = None,
) -> PollResultOut:
    """Poll device via SDK in an isolated subprocess (crash-safe)."""
    from datetime import datetime, timezone
    from cctv_monitor.polling.sdk_subprocess import poll_device_via_sdk

    result = await poll_device_via_sdk(
        host=device.host, port=device.sdk_port,
        username=device.username, password=password,
        lib_path=lib_path,
    )

    now = datetime.now(timezone.utc)

    if not result["success"]:
        health = HealthSummaryOut(
            reachable=False, camera_count=0, online_cameras=0,
            offline_cameras=0, disk_ok=False, response_time_ms=0,
            checked_at=now,
            web_port_open=web_port_open, sdk_port_open=sdk_port_open,
        )
        health_json_err = {
            "reachable": False, "camera_count": 0,
            "online_cameras": 0, "offline_cameras": 0,
            "disk_ok": False, "response_time_ms": 0,
            "checked_at": now.isoformat(),
            "web_port_open": web_port_open, "sdk_port_open": sdk_port_open,
        }
        try:
            health_log_repo = DeviceHealthLogRepository(session)
            await health_log_repo.insert(
                device_id=device_id, reachable=False,
                camera_count=0, online_cameras=0,
                offline_cameras=0, disk_ok=False, response_time_ms=0,
            )
            await repo.update(device_id, last_poll_at=now, last_health_json={
                "cameras": [], "disks": [], "health": health_json_err,
            })
            await session.commit()
        except Exception:
            pass
        return PollResultOut(device_id=device_id, health=health)

    # Parse subprocess results
    cameras = result.get("cameras", [])
    disks = result.get("disks", [])
    recordings = result.get("recordings", [])
    response_time = result.get("response_time_ms", 0)

    # Build recording map: channel_id -> recording status
    rec_map: dict[str, str] = {}
    for r in recordings:
        rec_map[r.get("channel_id", "")] = r.get("recording", "unknown")

    online = sum(1 for c in cameras if c.get("online") is True)
    disk_ok = all(d.get("status") in ("normal", "ok") for d in disks) if disks else True

    health = HealthSummaryOut(
        reachable=True, camera_count=len(cameras),
        online_cameras=online, offline_cameras=len(cameras) - online,
        disk_ok=disk_ok, response_time_ms=response_time,
        checked_at=now,
        web_port_open=web_port_open, sdk_port_open=sdk_port_open,
    )

    cameras_json = [
        {"channel_id": c.get("channel_id", ""), "channel_name": c.get("channel_name", ""),
         "status": "online" if c.get("online") else ("offline" if c.get("online") is False else "unknown"),
         "ip_address": c.get("ip_address"),
         "recording": rec_map.get(c.get("channel_id", "")),
         "checked_at": now.isoformat()}
        for c in cameras
    ]
    disks_json = [
        {"disk_id": str(d.get("disk_id", "")), "status": d.get("status", "unknown"),
         "capacity_bytes": d.get("capacity_bytes", 0), "free_bytes": d.get("free_bytes", 0),
         "health_status": d.get("health_status", "unknown"),
         "checked_at": now.isoformat(),
         "temperature": d.get("temperature"), "power_on_hours": d.get("power_on_hours"),
         "smart_status": d.get("smart_status")}
        for d in disks
    ]
    health_json = {
        "reachable": True, "camera_count": len(cameras),
        "online_cameras": online, "offline_cameras": len(cameras) - online,
        "disk_ok": disk_ok, "response_time_ms": response_time,
        "checked_at": now.isoformat(),
        "web_port_open": web_port_open, "sdk_port_open": sdk_port_open,
        "time_check": result.get("time_check"),
    }

    update_fields: dict = {
        "last_poll_at": now,
        "last_health_json": {
            "cameras": cameras_json, "disks": disks_json, "health": health_json,
        },
    }
    device_info = result.get("device_info")
    if device_info and isinstance(device_info, dict):
        if device_info.get("model"):
            update_fields["model"] = device_info["model"]
        if device_info.get("serial_number"):
            update_fields["serial_number"] = device_info["serial_number"]
        if device_info.get("firmware_version"):
            update_fields["firmware_version"] = device_info["firmware_version"]

    await repo.update(device_id, **update_fields)
    health_log_repo = DeviceHealthLogRepository(session)
    await health_log_repo.insert(
        device_id=device_id, reachable=True,
        camera_count=len(cameras), online_cameras=online,
        offline_cameras=len(cameras) - online,
        disk_ok=disk_ok, response_time_ms=response_time,
    )
    await session.commit()

    return PollResultOut(device_id=device_id, health=health)


@router.get("/devices/{device_id}/poll-stream")
async def poll_device_stream(
    device_id: str,
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
    registry: DriverRegistry = Depends(get_driver_registry),
    http_client: HttpClientManager = Depends(get_http_client),
):
    """SSE endpoint — streams poll steps in real time."""
    import json as _json
    import time as _time
    from datetime import datetime, timezone

    repo = DeviceRepository(session)
    device = await repo.get_by_id(device_id)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")

    password = decrypt_value(device.password_encrypted, settings.ENCRYPTION_KEY)

    async def _event_stream():
        def _sse(step: str, status: str, detail: str = "", **extra):
            data = {"step": step, "status": status, "detail": detail, **extra}
            return f"data: {_json.dumps(data, ensure_ascii=False)}\n\n"

        web_port_open = None
        sdk_port_open = None
        transport_used = None
        connect_error = ""  # diagnostic message for connection failure
        auth_failed = False  # set to True on auth error to prevent SDK fallback

        # Step 1: Check web port
        if device.web_port:
            yield _sse("web_port", "running", f"Checking TCP port {device.web_port}...")
            web_port_open = await _check_tcp_port(device.host, device.web_port, timeout=5.0)
            yield _sse("web_port", "success" if web_port_open else "error",
                        f"Port {device.web_port} {'open' if web_port_open else 'closed'}")
        else:
            yield _sse("web_port", "skipped", "Not configured")

        # Step 2: Check SDK port
        if device.sdk_port:
            yield _sse("sdk_port", "running", f"Checking TCP port {device.sdk_port}...")
            sdk_port_open = await _check_tcp_port(device.host, device.sdk_port, timeout=5.0)
            yield _sse("sdk_port", "success" if sdk_port_open else "error",
                        f"Port {device.sdk_port} {'open' if sdk_port_open else 'closed'}")
        else:
            yield _sse("sdk_port", "skipped", "Not configured")

        # --- Early exit: both ports unreachable ---
        both_ports_closed = (
            (device.web_port and web_port_open is False)
            and (device.sdk_port and sdk_port_open is False)
        )
        only_web_closed = (
            device.web_port and web_port_open is False
            and not device.sdk_port
        )
        only_sdk_closed = (
            device.sdk_port and sdk_port_open is False
            and not device.web_port
        )
        no_ports = not device.web_port and not device.sdk_port

        if both_ports_closed:
            connect_error = (
                f"Device unreachable — both ports closed: "
                f"web {device.host}:{device.web_port}, SDK {device.host}:{device.sdk_port}. "
                f"Check: IP address, network/internet, device power"
            )
        elif only_web_closed:
            connect_error = (
                f"Web port {device.host}:{device.web_port} closed. "
                f"Check: IP address, port number, network, device power"
            )
        elif only_sdk_closed:
            connect_error = (
                f"SDK port {device.host}:{device.sdk_port} closed. "
                f"Check: IP address, port number, network, device power"
            )
        elif no_ports:
            connect_error = "No ports configured — set web port or SDK port in device settings"

        if connect_error:
            yield _sse("connect", "error", connect_error)
            for skip_step in ("device_info", "cameras", "disks", "recording", "time_check"):
                yield _sse(skip_step, "skipped", "")
            yield _sse("done", "error", connect_error)
            # Save failure
            now = datetime.now(timezone.utc)
            health_json = {
                "reachable": False, "camera_count": 0,
                "online_cameras": 0, "offline_cameras": 0,
                "disk_ok": False, "response_time_ms": 0,
                "checked_at": now.isoformat(),
                "web_port_open": web_port_open, "sdk_port_open": sdk_port_open,
            }
            try:
                await repo.update(device_id, last_poll_at=now, last_health_json={
                    "cameras": [], "disks": [], "health": health_json,
                })
                health_log_repo = DeviceHealthLogRepository(session)
                await health_log_repo.insert(
                    device_id=device_id, reachable=False,
                    camera_count=0, online_cameras=0,
                    offline_cameras=0, disk_ok=False, response_time_ms=0,
                )
                await session.commit()
            except Exception:
                pass
            return

        # Step 3: Connect — choose transport
        driver = None
        connected = False

        if device.web_port and web_port_open:
            transport_used = "isapi"
            yield _sse("connect", "running", f"Connecting via ISAPI ({device.host}:{device.web_port})...")
            try:
                vendor = DeviceVendor(device.vendor)
                driver_cls = registry.get(vendor)
                transport_obj = IsapiTransport(http_client)
                driver = driver_cls(transport_obj)
                config = DeviceConfig(
                    device_id=device.device_id, name=device.name, vendor=vendor,
                    host=device.host, web_port=device.web_port, sdk_port=device.sdk_port,
                    web_protocol=device.web_protocol,
                    username=device.username, password=password,
                    transport_mode=DeviceTransport(device.transport_mode),
                    polling_policy_id=device.polling_policy_id, is_active=device.is_active,
                )
                await driver.connect(config, port=device.web_port)
                connected = True
                yield _sse("connect", "success", "ISAPI connected")
            except IsapiAuthError:
                auth_failed = True
                connect_error = (
                    f"Authentication failed — wrong username or password "
                    f"(ISAPI {device.host}:{device.web_port}). "
                    f"SDK fallback skipped to prevent device lockout."
                )
                yield _sse("connect", "error", connect_error)
                driver = None
            except Exception as exc:
                exc_msg = str(exc)
                if "ConnectTimeout" in exc_msg or "timed out" in exc_msg.lower():
                    connect_error = f"Connection timeout — {device.host}:{device.web_port} not responding"
                elif "ConnectError" in exc_msg or "ConnectionRefused" in exc_msg:
                    connect_error = f"Connection refused — {device.host}:{device.web_port}"
                else:
                    connect_error = f"ISAPI connection failed: {exc_msg}"
                yield _sse("connect", "error", connect_error)
                driver = None

        if not connected and not auth_failed and device.sdk_port and sdk_port_open:
            transport_used = "sdk"
            yield _sse("connect", "running", f"Connecting via SDK ({device.host}:{device.sdk_port})...")
            # SDK uses subprocess — we'll handle it below
            yield _sse("connect", "success", "Using SDK subprocess")
            connected = True

        if not connected:
            # Auth error or all transports failed
            if not connect_error:
                connect_error = "No transport available — all connection attempts failed"
            for skip_step in ("device_info", "cameras", "disks", "recording", "time_check"):
                yield _sse(skip_step, "skipped", "")
            yield _sse("done", "error", connect_error)
            # Save failure
            now = datetime.now(timezone.utc)
            health_json = {
                "reachable": False, "camera_count": 0,
                "online_cameras": 0, "offline_cameras": 0,
                "disk_ok": False, "response_time_ms": 0,
                "checked_at": now.isoformat(),
                "web_port_open": web_port_open, "sdk_port_open": sdk_port_open,
            }
            try:
                await repo.update(device_id, last_poll_at=now, last_health_json={
                    "cameras": [], "disks": [], "health": health_json,
                })
                health_log_repo = DeviceHealthLogRepository(session)
                await health_log_repo.insert(
                    device_id=device_id, reachable=False,
                    camera_count=0, online_cameras=0,
                    offline_cameras=0, disk_ok=False, response_time_ms=0,
                )
                await session.commit()
            except Exception:
                pass
            return

        # Ignored channels — exclude from all counts/reports
        ignored = {str(ch) for ch in (device.ignored_channels or [])}

        # ── ISAPI path ──
        if transport_used == "isapi" and driver is not None:
            start_t = _time.monotonic()

            # Step 4: Device info
            device_info = None
            yield _sse("device_info", "running", "Getting device info...")
            try:
                device_info = await driver.get_device_info()
                detail = f"{device_info.model}" if device_info and device_info.model else "OK"
                yield _sse("device_info", "success", detail)
            except Exception as exc:
                yield _sse("device_info", "error", str(exc))

            # Step 5: Cameras
            cameras_raw = []
            yield _sse("cameras", "running", "Getting camera statuses...")
            try:
                cameras_raw = await driver.get_camera_statuses()
                monitored = [c for c in cameras_raw if str(c.channel_id) not in ignored]
                online = sum(1 for c in monitored if c.status.value == "online")
                yield _sse("cameras", "success",
                           f"{len(monitored)} cameras, {online} online")
            except Exception as exc:
                yield _sse("cameras", "error", str(exc))

            # Step 6: Disks
            disks_raw = []
            yield _sse("disks", "running", "Getting disk statuses...")
            try:
                disks_raw = await driver.get_disk_statuses()
                all_ok = all(d.status.value == "ok" for d in disks_raw) if disks_raw else True
                smart_count = sum(1 for d in disks_raw if d.smart_status is not None)
                smart_info = f", SMART: {smart_count}/{len(disks_raw)}" if disks_raw else ""
                yield _sse("disks", "success",
                           f"{len(disks_raw)} disks, {'all OK' if all_ok else 'ERROR'}{smart_info}")
            except Exception as exc:
                yield _sse("disks", "error", str(exc))

            # Step 7: Recording status — prefer SDK FindFile over ISAPI search
            rec_map: dict[str, str] = {}
            yield _sse("recording", "running", "Checking recording status...")
            try:
                monitored_ids = [c.channel_id for c in cameras_raw if str(c.channel_id) not in ignored]
                if device.sdk_port and settings.HCNETSDK_LIB_PATH:
                    from cctv_monitor.polling.sdk_subprocess import poll_device_via_sdk_recordings
                    sdk_recs = await poll_device_via_sdk_recordings(
                        host=device.host, port=device.sdk_port,
                        username=device.username, password=password,
                        lib_path=settings.HCNETSDK_LIB_PATH,
                        channels=monitored_ids,
                    )
                    for r in sdk_recs:
                        ch = r.get("channel_id", "")
                        if str(ch) not in ignored:
                            rec_map[ch] = r.get("recording", "unknown")
                if not rec_map:
                    rec_statuses = await driver.get_recording_statuses()
                    for r in rec_statuses:
                        if str(r.channel_id) not in ignored:
                            rec_map[r.channel_id] = r.status.value
                recording_count = sum(1 for v in rec_map.values() if v == "recording")
                yield _sse("recording", "success",
                           f"{recording_count}/{len(rec_map)} recording" if rec_map else "Not available")
            except Exception as exc:
                yield _sse("recording", "error", str(exc))

            # Step 8: Time check
            time_check_data: dict | None = None
            yield _sse("time_check", "running", "Checking device time...")
            try:
                time_check_data = await driver.get_device_time()
                if time_check_data:
                    drift = time_check_data["drift_seconds"]
                    abs_drift = abs(drift)
                    if abs_drift < 5:
                        drift_label = "synced"
                    elif abs_drift < 30:
                        drift_label = f"{drift:+d}s"
                    elif abs_drift < 3600:
                        drift_label = f"{drift:+d}s ({abs_drift // 60}m)"
                    else:
                        drift_label = f"{drift:+d}s ({abs_drift // 3600}h {(abs_drift % 3600) // 60}m)"
                    mode = time_check_data.get("time_mode", "")
                    mode_label = f", {mode}" if mode else ""
                    status = "success" if abs_drift < 30 else ("error" if abs_drift > 300 else "success")
                    yield _sse("time_check", status, f"{drift_label}{mode_label}")
                else:
                    yield _sse("time_check", "skipped", "Not supported")
            except Exception as exc:
                yield _sse("time_check", "skipped", str(exc))

            response_time = (_time.monotonic() - start_t) * 1000
            now = datetime.now(timezone.utc)
            all_online = sum(1 for c in cameras_raw if c.status.value == "online")
            disk_ok = all(d.status.value == "ok" for d in disks_raw) if disks_raw else True

            # Save results (all cameras in JSON, ignored filtering at read time via _device_out)
            cameras_json = [
                {"channel_id": c.channel_id, "channel_name": c.channel_name,
                 "status": c.status.value, "ip_address": c.ip_address,
                 "recording": rec_map.get(c.channel_id),
                 "checked_at": c.checked_at.isoformat()}
                for c in cameras_raw
            ]
            disks_json = [
                {"disk_id": d.disk_id, "status": d.status.value,
                 "capacity_bytes": d.capacity_bytes, "free_bytes": d.free_bytes,
                 "health_status": d.health_status,
                 "checked_at": d.checked_at.isoformat(),
                 "temperature": d.temperature, "power_on_hours": d.power_on_hours,
                 "smart_status": d.smart_status}
                for d in disks_raw
            ]
            # Monitored counts (for SSE display and done message)
            monitored = [c for c in cameras_raw if str(c.channel_id) not in ignored]
            mon_online = sum(1 for c in monitored if c.status.value == "online")

            health_json = {
                "reachable": True, "camera_count": len(cameras_raw),
                "online_cameras": all_online, "offline_cameras": len(cameras_raw) - all_online,
                "disk_ok": disk_ok, "response_time_ms": response_time,
                "checked_at": now.isoformat(),
                "web_port_open": web_port_open, "sdk_port_open": sdk_port_open,
                "time_check": time_check_data,
            }
            update_fields: dict = {
                "last_poll_at": now,
                "last_health_json": {
                    "cameras": cameras_json, "disks": disks_json, "health": health_json,
                },
            }
            if device_info:
                update_fields["model"] = device_info.model
                update_fields["serial_number"] = device_info.serial_number
                update_fields["firmware_version"] = device_info.firmware_version

            await repo.update(device_id, **update_fields)
            health_log_repo = DeviceHealthLogRepository(session)
            await health_log_repo.insert(
                device_id=device_id, reachable=True,
                camera_count=len(cameras_raw), online_cameras=all_online,
                offline_cameras=len(cameras_raw) - all_online,
                disk_ok=disk_ok, response_time_ms=response_time,
            )
            await session.commit()

            # Evaluate alerts (resolve stale ones, create new)
            try:
                from cctv_monitor.alerts.engine import AlertEngine
                from cctv_monitor.models.device_health import DeviceHealthSummary
                from cctv_monitor.storage.repositories import AlertRepository
                from cctv_monitor.storage.tables import AlertTable
                from cctv_monitor.core.types import AlertStatus as _AS

                _eng = AlertEngine()
                _mon = [c for c in cameras_json if str(c.get("channel_id", "")) not in ignored]
                _rec_t = sum(1 for c in _mon if c.get("recording"))
                _rec_ok = sum(1 for c in _mon if c.get("recording") == "recording")
                _summary = DeviceHealthSummary(
                    device_id=device_id, reachable=True,
                    camera_count=len(monitored), online_cameras=mon_online,
                    offline_cameras=len(monitored) - mon_online,
                    disk_ok=disk_ok,
                    recording_ok=(_rec_ok == _rec_t) if _rec_t > 0 else True,
                    response_time_ms=response_time,
                    checked_at=now,
                )
                _alert_repo = AlertRepository(session)
                _active = await _alert_repo.get_active_alerts(device_id)
                from cctv_monitor.models.alert import AlertEvent as _AE
                _active_events = [
                    _AE(device_id=r.device_id, alert_type=r.alert_type, severity=r.severity,
                        message=r.message, source=r.source, status=r.status,
                        created_at=r.created_at, id=r.id, channel_id=r.channel_id, resolved_at=r.resolved_at)
                    for r in _active
                ]
                _new, _resolved = _eng.evaluate(_summary, _active_events)
                for a in _new:
                    await _alert_repo.create_alert(AlertTable(
                        device_id=a.device_id, alert_type=a.alert_type, severity=a.severity,
                        message=a.message, source=a.source, status=_AS.ACTIVE, created_at=a.created_at,
                    ))
                for a in _resolved:
                    if a.id is not None:
                        await _alert_repo.resolve_alert(a.id)
                await session.commit()
            except Exception:
                pass

            try:
                await driver.disconnect()
            except Exception:
                pass

            yield _sse("done", "success",
                        f"Poll complete — {len(monitored)} cameras, {mon_online} online, "
                        f"disks {'OK' if disk_ok else 'ERROR'}, {round(response_time)}ms",
                        response_time_ms=round(response_time, 1))

        # ── SDK path ──
        elif transport_used == "sdk":
            from cctv_monitor.polling.sdk_subprocess import poll_device_via_sdk

            yield _sse("device_info", "running", "Polling via SDK subprocess...")

            result = await poll_device_via_sdk(
                host=device.host, port=device.sdk_port,
                username=device.username, password=password,
                lib_path=settings.HCNETSDK_LIB_PATH,
            )

            if not result["success"]:
                raw_err = result.get("error", "SDK failed")
                # Parse SDK error codes for user-friendly messages
                if "error 1:" in raw_err.lower() or "error_code=1" in raw_err:
                    err_msg = (
                        f"Authentication failed — wrong username or password "
                        f"(SDK {device.host}:{device.sdk_port})"
                    )
                elif "error 3:" in raw_err.lower():
                    err_msg = f"SDK initialization error — {raw_err}"
                elif "error 7:" in raw_err.lower():
                    err_msg = (
                        f"Network error connecting to {device.host}:{device.sdk_port} — "
                        f"check IP address and port"
                    )
                elif "not reachable" in raw_err.lower():
                    err_msg = f"SDK port {device.host}:{device.sdk_port} is not reachable"
                elif "timeout" in raw_err.lower():
                    err_msg = f"SDK connection timeout — {device.host}:{device.sdk_port} not responding"
                else:
                    err_msg = f"SDK connection failed: {raw_err}"
                yield _sse("device_info", "error", err_msg)
                for skip_step in ("cameras", "disks", "recording", "time_check"):
                    yield _sse(skip_step, "skipped", "")

                # Save failure
                now = datetime.now(timezone.utc)
                sdk_port_open = sdk_port_open if sdk_port_open is not None else False
                health_json = {
                    "reachable": False, "camera_count": 0,
                    "online_cameras": 0, "offline_cameras": 0,
                    "disk_ok": False, "response_time_ms": 0,
                    "checked_at": now.isoformat(),
                    "web_port_open": web_port_open, "sdk_port_open": sdk_port_open,
                }
                try:
                    await repo.update(device_id, last_poll_at=now, last_health_json={
                        "cameras": [], "disks": [], "health": health_json,
                    })
                    health_log_repo = DeviceHealthLogRepository(session)
                    await health_log_repo.insert(
                        device_id=device_id, reachable=False,
                        camera_count=0, online_cameras=0,
                        offline_cameras=0, disk_ok=False, response_time_ms=0,
                    )
                    await session.commit()
                except Exception:
                    pass

                yield _sse("done", "error", f"SDK poll failed: {err_msg}")
                return

            # SDK success
            dev_info = result.get("device_info")
            if dev_info and isinstance(dev_info, dict) and dev_info.get("model"):
                yield _sse("device_info", "success", dev_info["model"])
            else:
                yield _sse("device_info", "success", "Connected")

            cameras = result.get("cameras", [])
            monitored_cams = [c for c in cameras if str(c.get("channel_id", "")) not in ignored]
            online = sum(1 for c in monitored_cams if c.get("online") is True)
            yield _sse("cameras", "success", f"{len(monitored_cams)} cameras, {online} online")

            disks = result.get("disks", [])
            disk_ok = all(d.get("status") in ("normal", "ok") for d in disks) if disks else True
            smart_count = sum(1 for d in disks if d.get("smart_status") is not None)
            smart_info = f", SMART: {smart_count}/{len(disks)}" if disks else ""
            yield _sse("disks", "success",
                        f"{len(disks)} disks, {'all OK' if disk_ok else 'ERROR'}{smart_info}")

            # Recording status from SDK worker
            recordings = result.get("recordings", [])
            rec_map: dict[str, str] = {}
            for r in recordings:
                ch = r.get("channel_id", "")
                if str(ch) not in ignored:
                    rec_map[ch] = r.get("recording", "unknown")
            if rec_map:
                recording_count = sum(1 for v in rec_map.values() if v == "recording")
                yield _sse("recording", "success", f"{recording_count}/{len(rec_map)} recording")
            else:
                yield _sse("recording", "skipped", "Not available via SDK")

            # Time check from SDK worker
            time_check_data = result.get("time_check")
            if time_check_data:
                drift = time_check_data.get("drift_seconds", 0)
                abs_drift = abs(drift)
                if abs_drift < 5:
                    drift_label = "synced"
                elif abs_drift < 30:
                    drift_label = f"{drift:+d}s"
                elif abs_drift < 3600:
                    drift_label = f"{drift:+d}s ({abs_drift // 60}m)"
                else:
                    drift_label = f"{drift:+d}s ({abs_drift // 3600}h {(abs_drift % 3600) // 60}m)"
                mode = time_check_data.get("time_mode", "")
                mode_label = f", {mode}" if mode else ""
                status = "success" if abs_drift < 30 else ("error" if abs_drift > 300 else "success")
                yield _sse("time_check", status, f"{drift_label}{mode_label}")
            else:
                yield _sse("time_check", "skipped", "Not available")

            now = datetime.now(timezone.utc)
            response_time = result.get("response_time_ms", 0)
            sdk_port_open = True  # SDK worked

            cameras_json = [
                {"channel_id": c.get("channel_id", ""), "channel_name": c.get("channel_name", ""),
                 "status": "online" if c.get("online") else ("offline" if c.get("online") is False else "unknown"),
                 "ip_address": c.get("ip_address"),
                 "recording": rec_map.get(c.get("channel_id", "")),
                 "checked_at": now.isoformat()}
                for c in cameras
            ]
            disks_json = [
                {"disk_id": str(d.get("disk_id", "")), "status": d.get("status", "unknown"),
                 "capacity_bytes": d.get("capacity_bytes", 0), "free_bytes": d.get("free_bytes", 0),
                 "health_status": d.get("health_status", "unknown"),
                 "checked_at": now.isoformat(),
                 "temperature": d.get("temperature"), "power_on_hours": d.get("power_on_hours"),
                 "smart_status": d.get("smart_status")}
                for d in disks
            ]
            all_online_sdk = sum(1 for c in cameras if c.get("online") is True)
            health_json = {
                "reachable": True, "camera_count": len(cameras),
                "online_cameras": all_online_sdk, "offline_cameras": len(cameras) - all_online_sdk,
                "disk_ok": disk_ok, "response_time_ms": response_time,
                "checked_at": now.isoformat(),
                "web_port_open": web_port_open, "sdk_port_open": sdk_port_open,
                "time_check": time_check_data,
            }
            update_fields: dict = {
                "last_poll_at": now,
                "last_health_json": {
                    "cameras": cameras_json, "disks": disks_json, "health": health_json,
                },
            }
            if dev_info and isinstance(dev_info, dict):
                if dev_info.get("model"):
                    update_fields["model"] = dev_info["model"]
                if dev_info.get("serial_number"):
                    update_fields["serial_number"] = dev_info["serial_number"]
                if dev_info.get("firmware_version"):
                    update_fields["firmware_version"] = dev_info["firmware_version"]

            await repo.update(device_id, **update_fields)
            health_log_repo = DeviceHealthLogRepository(session)
            await health_log_repo.insert(
                device_id=device_id, reachable=True,
                camera_count=len(cameras), online_cameras=all_online_sdk,
                offline_cameras=len(cameras) - all_online_sdk,
                disk_ok=disk_ok, response_time_ms=response_time,
            )
            await session.commit()

            # Evaluate alerts (resolve stale ones, create new)
            try:
                from cctv_monitor.alerts.engine import AlertEngine
                from cctv_monitor.models.device_health import DeviceHealthSummary
                from cctv_monitor.storage.repositories import AlertRepository
                from cctv_monitor.storage.tables import AlertTable
                from cctv_monitor.core.types import AlertStatus as _AS

                _eng = AlertEngine()
                _mon = [c for c in cameras_json if str(c.get("channel_id", "")) not in ignored]
                _rec_t = sum(1 for c in _mon if c.get("recording"))
                _rec_ok = sum(1 for c in _mon if c.get("recording") == "recording")
                _summary = DeviceHealthSummary(
                    device_id=device_id, reachable=True,
                    camera_count=len(monitored_cams), online_cameras=online,
                    offline_cameras=len(monitored_cams) - online,
                    disk_ok=disk_ok,
                    recording_ok=(_rec_ok == _rec_t) if _rec_t > 0 else True,
                    response_time_ms=response_time,
                    checked_at=now,
                )
                _alert_repo = AlertRepository(session)
                _active = await _alert_repo.get_active_alerts(device_id)
                from cctv_monitor.models.alert import AlertEvent as _AE
                _active_events = [
                    _AE(device_id=r.device_id, alert_type=r.alert_type, severity=r.severity,
                        message=r.message, source=r.source, status=r.status,
                        created_at=r.created_at, id=r.id, channel_id=r.channel_id, resolved_at=r.resolved_at)
                    for r in _active
                ]
                _new, _resolved = _eng.evaluate(_summary, _active_events)
                for a in _new:
                    await _alert_repo.create_alert(AlertTable(
                        device_id=a.device_id, alert_type=a.alert_type, severity=a.severity,
                        message=a.message, source=a.source, status=_AS.ACTIVE, created_at=a.created_at,
                    ))
                for a in _resolved:
                    if a.id is not None:
                        await _alert_repo.resolve_alert(a.id)
                await session.commit()
            except Exception:
                pass

            yield _sse("done", "success",
                        f"Poll complete — {len(monitored_cams)} cameras, {online} online, "
                        f"disks {'OK' if disk_ok else 'ERROR'}, {round(response_time)}ms",
                        response_time_ms=round(response_time, 1))

    return StreamingResponse(
        _event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# Limit concurrent snapshot requests per device to avoid overwhelming NVRs
_snapshot_semaphores: dict[str, asyncio.Semaphore] = {}
_SNAPSHOT_CONCURRENCY = 1  # keep old NVRs stable under snapshot storms


def _get_snapshot_semaphore(device_id: str) -> asyncio.Semaphore:
    if device_id not in _snapshot_semaphores:
        _snapshot_semaphores[device_id] = asyncio.Semaphore(_SNAPSHOT_CONCURRENCY)
    return _snapshot_semaphores[device_id]


# SDK batch snapshot cache: one subprocess captures all channels, results cached briefly
import time as _time

_sdk_batch_cache: dict[str, tuple[float, dict[str, bytes], dict[str, str]]] = {}
_sdk_batch_locks: dict[str, asyncio.Lock] = {}
_SDK_BATCH_TTL = 25  # seconds — slightly less than frontend refresh interval (30s)
_sdk_batch_disabled_until: dict[str, float] = {}
_sdk_batch_timeout_streak: dict[str, int] = {}
_SDK_BATCH_TIMEOUT_STREAK_LIMIT = 2
_SDK_BATCH_DISABLE_SECONDS = 300


def _get_sdk_batch_lock(device_id: str) -> asyncio.Lock:
    if device_id not in _sdk_batch_locks:
        _sdk_batch_locks[device_id] = asyncio.Lock()
    return _sdk_batch_locks[device_id]


@router.get("/devices/{device_id}/snapshots")
async def get_snapshots_page(
    device_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(16, ge=1, le=32),
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
):
    """Return base64 JPEG snapshots for one page of channels."""
    import base64 as _b64

    repo = DeviceRepository(session)
    device = await repo.get_by_id(device_id)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")

    # Build channel list (excluding ignored)
    all_channels = _get_channel_ids_from_health(device.last_health_json)
    ignored = set(device.ignored_channels or [])
    channels = [ch for ch in all_channels if ch not in ignored]

    total = len(channels)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = min(page, total_pages)
    start = (page - 1) * page_size
    page_channels = channels[start : start + page_size]

    if not page_channels:
        return {
            "snapshots": {},
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": total_pages,
        }

    password = decrypt_value(device.password_encrypted, settings.ENCRYPTION_KEY)
    sdk_available = bool(device.sdk_port and settings.HCNETSDK_LIB_PATH)
    snapshots: dict[str, str | None] = {}

    semaphore = _get_snapshot_semaphore(device_id)
    async with semaphore:
        if sdk_available:
            results, errors = await _sdk_batch_snapshot_subprocess(
                device.host, device.sdk_port, device.username, password,
                page_channels, settings.HCNETSDK_LIB_PATH,
            )
            for ch in page_channels:
                img = results.get(ch)
                snapshots[ch] = _b64.b64encode(img).decode() if img else None
        else:
            # Non-SDK: return None per channel, frontend uses individual snapshot URLs
            for ch in page_channels:
                snapshots[ch] = None

    return {
        "snapshots": snapshots,
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": total_pages,
    }


@router.get("/devices/{device_id}/snapshot/{channel_id}")
async def get_snapshot(
    device_id: str,
    channel_id: str,
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
    http_client: HttpClientManager = Depends(get_http_client),
):
    repo = DeviceRepository(session)
    device = await repo.get_by_id(device_id)
    if device is None:
        raise HTTPException(status_code=404, detail="Device not found")

    password = decrypt_value(device.password_encrypted, settings.ENCRYPTION_KEY)
    sdk_available = bool(device.sdk_port and settings.HCNETSDK_LIB_PATH)
    prefer_sdk = device.transport_mode == DeviceTransport.SDK
    if not sdk_available:
        logger.info(
            "snapshot: SDK not available for device=%s (sdk_port=%s, lib_path=%s)",
            device_id, device.sdk_port, bool(settings.HCNETSDK_LIB_PATH),
        )

    semaphore = _get_snapshot_semaphore(device_id)
    try:
        async with semaphore:
            image_data = None
            errors: list[str] = []

            # Try ISAPI first (unless SDK-preferred)
            if not prefer_sdk and device.web_port:
                transport = IsapiTransport(http_client)
                try:
                    await transport.connect(device.host, device.web_port, device.username, password, protocol=device.web_protocol)
                    image_data = await transport.get_snapshot(channel_id)
                except IsapiAuthError as auth_exc:
                    errors.append(f"ISAPI: {auth_exc}")
                    logger.warning("ISAPI snapshot auth failed for device=%s channel=%s", device_id, channel_id)
                    raise HTTPException(status_code=502, detail=f"Auth failed: {auth_exc}")
                except Exception as isapi_exc:
                    errors.append(f"ISAPI: {isapi_exc}")
                    logger.warning("ISAPI snapshot failed for device=%s channel=%s: %s", device_id, channel_id, isapi_exc)
                finally:
                    try:
                        await transport.disconnect()
                    except Exception:
                        pass

            # SDK fallback (or primary if SDK-preferred): use batch cache
            if image_data is None and sdk_available:
                image_data = await _get_sdk_snapshot_cached(
                    device_id, device.host, device.sdk_port,
                    device.username, password, channel_id, settings.HCNETSDK_LIB_PATH,
                    health_json=device.last_health_json,
                )
                if image_data is None:
                    sdk_err = _sdk_batch_cache.get(device_id)
                    if sdk_err:
                        _, _, ch_errors = sdk_err
                        err = ch_errors.get(channel_id, "unknown error")
                        errors.append(f"SDK: {err}")

            if image_data is None:
                if not device.web_port and not sdk_available:
                    raise HTTPException(status_code=400, detail="No port configured for snapshots")
                raise HTTPException(status_code=502, detail=f"All snapshot methods failed: {'; '.join(errors)}")
    except HTTPException:
        raise
    except Exception as exc:
        logger.warning("Snapshot failed for device=%s channel=%s: %s", device_id, channel_id, exc)
        raise HTTPException(status_code=502, detail=f"Failed to get snapshot: {exc}")

    return StreamingResponse(
        io.BytesIO(image_data),
        media_type="image/jpeg",
        headers={"Cache-Control": "public, max-age=30"},
    )


async def _sdk_snapshot_subprocess(
    host: str, port: int, username: str, password: str,
    channel: int, lib_path: str,
) -> bytes:
    """Run SDK snapshot capture in an isolated subprocess."""
    import sys

    cmd = [
        sys.executable, "-m", "cctv_monitor.polling.sdk_worker",
        "--host", host,
        "--port", str(port),
        "--user", username,
        "--password", password,
        "--lib-path", lib_path,
        "--snapshot",
        "--channel", str(channel),
    ]

    proc = await asyncio.create_subprocess_exec(
        *cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )

    try:
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=35)
    except asyncio.TimeoutError:
        proc.kill()
        await proc.wait()
        raise Exception("SDK snapshot subprocess timed out")

    # Log subprocess debug output (channel mapping info etc.)
    if stderr:
        debug_info = stderr.decode(errors="replace").strip()
        if debug_info:
            logger.info("SDK snapshot subprocess ch=%s: %s", channel, debug_info)

    if proc.returncode != 0:
        err_msg = stderr.decode(errors="replace").strip() if stderr else ""
        if not err_msg:
            # Crashed subprocess (e.g. 0xC0000005 = access violation)
            err_msg = f"subprocess crashed (exit code {proc.returncode})"
        raise Exception(f"SDK snapshot failed: {err_msg}")

    if not stdout or len(stdout) < 2:
        raise Exception("SDK snapshot returned empty data")

    # Validate JPEG header
    if stdout[:2] != b"\xff\xd8":
        raise Exception("SDK snapshot returned invalid JPEG data")

    return stdout


async def _sdk_single_snapshot_fallback(
    *,
    device_id: str,
    host: str,
    port: int,
    username: str,
    password: str,
    channel_id: str,
    lib_path: str,
) -> bytes | None:
    try:
        ch_int = int(channel_id)
    except (TypeError, ValueError):
        return None
    try:
        logger.info(
            "SDK single snapshot fallback for device=%s channel=%s",
            device_id, channel_id,
        )
        return await _sdk_snapshot_subprocess(
            host, port, username, password, ch_int, lib_path,
        )
    except Exception as single_exc:
        logger.warning(
            "SDK single snapshot fallback failed for device=%s channel=%s: %s",
            device_id, channel_id, single_exc,
        )
        return None


async def _get_sdk_snapshot_cached(
    device_id: str, host: str, port: int, username: str, password: str,
    channel_id: str, lib_path: str, health_json: dict | None = None,
) -> bytes | None:
    """Get SDK snapshot from batch cache, triggering batch capture if needed.

    When the first channel request arrives, we launch ONE subprocess that captures
    ALL channels for the device. Subsequent requests within TTL read from cache.
    """
    now = _time.time()
    if _sdk_batch_disabled_until.get(device_id, 0.0) > now:
        return await _sdk_single_snapshot_fallback(
            device_id=device_id,
            host=host,
            port=port,
            username=username,
            password=password,
            channel_id=channel_id,
            lib_path=lib_path,
        )

    # Check cache first
    cached = _sdk_batch_cache.get(device_id)
    if cached:
        ts, results, errors = cached
        if _time.time() - ts < _SDK_BATCH_TTL:
            image = results.get(channel_id)
            if image is not None:
                return image
            image = await _sdk_single_snapshot_fallback(
                device_id=device_id,
                host=host,
                port=port,
                username=username,
                password=password,
                channel_id=channel_id,
                lib_path=lib_path,
            )
            if image is not None:
                results[channel_id] = image
                _sdk_batch_cache[device_id] = (ts, results, errors)
            return image

    # Need to fetch — acquire per-device lock so only one batch runs
    lock = _get_sdk_batch_lock(device_id)
    async with lock:
        # Double-check after acquiring lock
        cached = _sdk_batch_cache.get(device_id)
        if cached:
            ts, results, errors = cached
            if _time.time() - ts < _SDK_BATCH_TTL:
                image = results.get(channel_id)
                if image is not None:
                    return image
                image = await _sdk_single_snapshot_fallback(
                    device_id=device_id,
                    host=host,
                    port=port,
                    username=username,
                    password=password,
                    channel_id=channel_id,
                    lib_path=lib_path,
                )
                if image is not None:
                    results[channel_id] = image
                    _sdk_batch_cache[device_id] = (ts, results, errors)
                return image

        # Get all channel IDs for this device from health data
        all_channels = _get_channel_ids_from_health(health_json)
        if not all_channels:
            all_channels = [channel_id]  # fallback: at least this channel

        logger.info(
            "SDK batch snapshot for device=%s channels=%s",
            device_id, ",".join(all_channels),
        )

        results, errors = await _sdk_batch_snapshot_subprocess(
            host, port, username, password, all_channels, lib_path,
        )

        _sdk_batch_cache[device_id] = (_time.time(), results, errors)

        if all_channels and len(results) == 0 and all(errors.get(ch) == "timeout" for ch in all_channels):
            streak = _sdk_batch_timeout_streak.get(device_id, 0) + 1
            _sdk_batch_timeout_streak[device_id] = streak
            if streak >= _SDK_BATCH_TIMEOUT_STREAK_LIMIT:
                _sdk_batch_disabled_until[device_id] = _time.time() + _SDK_BATCH_DISABLE_SECONDS
                logger.warning(
                    "SDK batch disabled temporarily for device=%s for %ss after %d timeout batches",
                    device_id, _SDK_BATCH_DISABLE_SECONDS, streak,
                )
        else:
            _sdk_batch_timeout_streak[device_id] = 0

        if errors:
            failed = list(errors.keys())[:3]
            logger.warning(
                "SDK batch snapshot errors for device=%s: %s (and %d more)",
                device_id, "; ".join(f"ch{k}: {errors[k]}" for k in failed),
                max(0, len(errors) - 3),
            )

        image = results.get(channel_id)
        if image is not None:
            return image

        # Batch can timeout on busy NVRs (many channels). Fallback to a
        # single-channel SDK capture so user still gets the requested frame.
        image = await _sdk_single_snapshot_fallback(
            device_id=device_id,
            host=host,
            port=port,
            username=username,
            password=password,
            channel_id=channel_id,
            lib_path=lib_path,
        )
        if image is not None:
            # Warm the cache with the successful single capture.
            cached_ts, cached_results, cached_errors = _sdk_batch_cache.get(
                device_id, (_time.time(), {}, {}),
            )
            cached_results[channel_id] = image
            _sdk_batch_cache[device_id] = (cached_ts, cached_results, cached_errors)
        return image


def _get_channel_ids_from_health(health_json: dict | None) -> list[str]:
    """Extract channel IDs from device health_json."""
    if not health_json or not isinstance(health_json, dict):
        return []
    cameras = health_json.get("cameras", [])
    return [str(c.get("channel_id", "")) for c in cameras if c.get("channel_id")]


async def _sdk_batch_snapshot_subprocess(
    host: str, port: int, username: str, password: str,
    channel_ids: list[str], lib_path: str,
) -> tuple[dict[str, bytes], dict[str, str]]:
    """Run batch SDK snapshot in one subprocess — one login, all channels."""
    import base64 as _b64
    import sys

    cmd = [
        sys.executable, "-m", "cctv_monitor.polling.sdk_worker",
        "--host", host,
        "--port", str(port),
        "--user", username,
        "--password", password,
        "--lib-path", lib_path,
        "--snapshot-batch",
        "--snapshot-channels", ",".join(channel_ids),
    ]

    proc = await asyncio.create_subprocess_exec(
        *cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )

    timeout_s = max(30, min(120, 5 * len(channel_ids)))
    try:
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout_s)
    except asyncio.TimeoutError:
        proc.kill()
        await proc.wait()
        return {}, {ch: "timeout" for ch in channel_ids}

    if stderr:
        debug_info = stderr.decode(errors="replace").strip()
        if debug_info:
            logger.info("SDK batch snapshot subprocess: %s", debug_info)

    if proc.returncode != 0:
        err_msg = stderr.decode(errors="replace").strip() if stderr else f"exit code {proc.returncode}"
        return {}, {ch: err_msg for ch in channel_ids}

    try:
        import json
        data = json.loads(stdout.decode("utf-8"))
        results: dict[str, bytes] = {}
        for ch_id, b64_jpeg in data.get("results", {}).items():
            jpeg = _b64.b64decode(b64_jpeg)
            if len(jpeg) >= 2 and jpeg[:2] == b"\xff\xd8":
                results[ch_id] = jpeg
        return results, data.get("errors", {})
    except Exception as exc:
        logger.warning("Failed to parse SDK batch snapshot output: %s", exc)
        return {}, {ch: str(exc) for ch in channel_ids}
